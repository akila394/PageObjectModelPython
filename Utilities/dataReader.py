import openpyxl

def get_data(datasheet):
    workbook = openpyxl.load_workbook("../Excelfiles//testdata.xlsx")
    sheet = workbook[datasheet]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainlist = []
    for i in range(2, totalrows + 1):
        datalist = []
        for j in range(1, totalcols + 1):
            data = sheet.cell(row = i, column= j).value
            datalist.insert(j, data)
        mainlist.insert(i, datalist)
    return mainlist
